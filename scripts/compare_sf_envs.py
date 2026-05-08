"""
Compare Salesforce prod vs nonprod responses for one or more ASURITEs.

Usage:
    python scripts/compare_sf_envs.py jsmith1 jdoe2 ...
    python scripts/compare_sf_envs.py --all          # reads every asurite from the DB

Output: a per-asurite report of any fields that differ between environments,
plus a summary of which ASURITEs matched and which diverged.
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

ENV_PATH = PROJECT_ROOT / ".env"

# asu_discord/__init__.py imports the discord voice stack which requires audioop
# (removed in Python 3.13+). Stub the package so salesforce.py loads cleanly.
import types as _types
_pkg = _types.ModuleType("asu_discord")
_pkg.__path__ = [str(PROJECT_ROOT / "asu_discord")]
_pkg.__package__ = "asu_discord"
sys.modules.setdefault("asu_discord", _pkg)

# Fields checked in the comparison — these are the ones that drive Discord role assignment.
COMPARED_FIELDS = [
    "name",
    "email",
    "college",
    "program",
    "career",
    "admitType",
    "firstTimeFreshman",
    "transfer",
    "inState",
    "outOfState",
    "international",
    "campus",
    "locationName",
    "termCode",
    "stageName",
    "current",
    "firstYear",
    "depositPaid",
    "enrollmentDepositPaid",
    "collegeProgramCode",
    "is_international",
]


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip("\"'")
        if key:
            env[key] = value
    return env


def apply_env(env: dict[str, str]) -> None:
    for key, value in env.items():
        os.environ.setdefault(key, value)


ESB_BASE = "https://esb.asu.edu"
ESB_CONTACT_URL = f"{ESB_BASE}/api/v1/asu-sf-contact/contact"
ESB_OPP_URL = f"{ESB_BASE}/api/v1/asu-sf-opportunity/opportunity"


def _profile_to_dict(profile) -> dict[str, Any]:
    if profile is None:
        return {}
    try:
        return profile.model_dump()
    except AttributeError:
        return vars(profile)


def _get_prod_profile_esb(asurite: str, client_id: str, client_secret: str):
    """Replicate master-branch Basic Auth logic against the ESB endpoint."""
    import base64
    import requests as _req
    from asu_discord.salesforce import (
        StudentProfile, _select_opportunity, _deposit_paid_from_opportunity,
        _parse_bool, _created_date, PROGRAM_CODE_TO_COLLEGE_ROLE,
    )

    token = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
    headers = {"Authorization": f"Basic {token}"}

    try:
        contact_resp = _req.get(f"{ESB_CONTACT_URL}?asurite={asurite}", headers=headers, timeout=10)
    except _req.RequestException as exc:
        logger.warning("ESB contact request failed for %s: %s", asurite, exc)
        return None

    contact_resp.raise_for_status()
    contact_data = contact_resp.json()

    if "contact" not in contact_data or not contact_data["contact"]:
        logger.info("No ESB contact found for %s", asurite)
        return None

    contact = contact_data["contact"][0]
    contact_id = contact.get("id")
    first_name = contact.get("firstname", "") or ""
    last_name = contact.get("lastname", "") or ""
    full_name = f"{first_name} {last_name}".strip()

    opportunities: list[dict] = []
    for career in ["Undergraduate", "Graduate"]:
        try:
            opp_resp = _req.get(
                f"{ESB_OPP_URL}?contactId={contact_id}&career={career}",
                headers=headers,
                timeout=10,
            )
        except _req.RequestException:
            continue
        if opp_resp.status_code == 200:
            opp_data = opp_resp.json()
            if isinstance(opp_data, list):
                opportunities.extend(opp_data)

    opportunities.sort(key=_created_date)
    selected_opp, is_current_student = _select_opportunity(opportunities)

    state = contact.get("state")
    country = contact.get("country")
    state_normalized = (state or "").strip().lower()

    profile: dict = {
        "asurite": asurite,
        "name": full_name,
        "fullName": full_name,
        "email": contact.get("email"),
        "state": state,
        "country": country,
        "contactId": contact_id,
        "opportunities": opportunities,
        "inState": state_normalized in {"arizona", "az"},
        "outOfState": bool(state_normalized and state_normalized not in {"arizona", "az"}),
    }

    if selected_opp is not None:
        college_program_code = selected_opp.get("collegeProgramCode") or selected_opp.get("programCode")
        is_international = _parse_bool(selected_opp.get("internationalStudent"))
        admit_type = selected_opp.get("type")
        career = selected_opp.get("career")
        in_state = (state_normalized in {"arizona", "az"}) if not is_international else False
        out_of_state = bool(not in_state and state_normalized) if not is_international else False

        profile.update({
            "college": PROGRAM_CODE_TO_COLLEGE_ROLE.get(college_program_code, selected_opp.get("collegeName") or "N/A"),
            "program": selected_opp.get("academicPlanName") or "None",
            "career": career or "Unknown",
            "admitType": admit_type or "None",
            "firstTimeFreshman": bool(isinstance(admit_type, str) and "freshman" in admit_type.strip().lower()),
            "transfer": bool(isinstance(admit_type, str) and "transfer" in admit_type.strip().lower()),
            "collegeProgramCode": college_program_code,
            "is_international": is_international,
            "stateOfResidence": state if not is_international else None,
            "enrollmentDepositPaid": _deposit_paid_from_opportunity(selected_opp),
            "stageName": selected_opp.get("stageName"),
            "current": is_current_student,
            "firstYear": False,
            "depositPaid": _deposit_paid_from_opportunity(selected_opp),
            "international": is_international,
            "inState": in_state,
            "outOfState": out_of_state,
            "campus": selected_opp.get("currentLocation") or "N/A",
            "locationName": selected_opp.get("locationName") or selected_opp.get("currentLocation") or "N/A",
            "termCode": selected_opp.get("termCode"),
            "selectedOpportunity": selected_opp,
        })
        if career and career.lower() == "undergraduate":
            if isinstance(admit_type, str) and "transfer" in admit_type.lower():
                profile.update({"firstYear": False, "transfer": True, "type": "Transfer"})
            elif isinstance(admit_type, str) and "freshman" in admit_type.lower():
                profile.update({"firstYear": True, "transfer": False, "type": "First Time Freshman"})
        elif career and career.lower() == "graduate":
            profile.update({"firstYear": False, "transfer": False, "type": "Masters"})

    if not opportunities:
        logger.info("No ESB opportunities found for %s", asurite)
        return None

    return StudentProfile.model_validate(profile)


def _safe_get_profile(asurite: str, creds: dict, label: str):
    import requests as _requests
    from asu_discord.salesforce import get_student_profile
    try:
        return get_student_profile(asurite, **creds), None
    except _requests.exceptions.HTTPError as exc:
        return None, f"HTTP {exc.response.status_code} from {label}: {exc}"
    except Exception as exc:
        return None, f"Error from {label}: {exc}"


def _safe_get_prod_esb(asurite: str, client_id: str, client_secret: str):
    import requests as _requests
    try:
        return _get_prod_profile_esb(asurite, client_id, client_secret), None
    except _requests.exceptions.HTTPError as exc:
        return None, f"HTTP {exc.response.status_code} from prod/ESB: {exc}"
    except Exception as exc:
        return None, f"Error from prod/ESB: {exc}"


def compare_asurite(asurite: str, prod_creds: dict, nonprod_creds: dict) -> dict:
    prod_profile, prod_err = _safe_get_prod_esb(asurite, prod_creds["client_id"], prod_creds["client_secret"])
    nonprod_profile, nonprod_err = _safe_get_profile(asurite, nonprod_creds, "nonprod")

    prod_data = _profile_to_dict(prod_profile)
    nonprod_data = _profile_to_dict(nonprod_profile)

    diffs: dict[str, dict] = {}
    for field in COMPARED_FIELDS:
        prod_val = prod_data.get(field)
        nonprod_val = nonprod_data.get(field)
        if prod_val != nonprod_val:
            diffs[field] = {"prod": prod_val, "nonprod": nonprod_val}

    return {
        "asurite": asurite,
        "prod_found": prod_profile is not None,
        "nonprod_found": nonprod_profile is not None,
        "prod_err": prod_err,
        "nonprod_err": nonprod_err,
        "diffs": diffs,
        "_prod_profile": prod_profile,
        "_nonprod_profile": nonprod_profile,
    }


def print_result(result: dict) -> None:
    asurite = result["asurite"]
    prod_err = result.get("prod_err")
    nonprod_err = result.get("nonprod_err")

    if prod_err:
        print(f"[{asurite}] PROD ERROR: {prod_err}")
        return
    if nonprod_err:
        print(f"[{asurite}] NONPROD ERROR: {nonprod_err}")
        return

    if not result["prod_found"] and not result["nonprod_found"]:
        print(f"[{asurite}] NOT FOUND in either environment")
        return
    if not result["prod_found"]:
        print(f"[{asurite}] NOT FOUND in prod (found in nonprod)")
        return
    if not result["nonprod_found"]:
        print(f"[{asurite}] NOT FOUND in nonprod (found in prod)")
        return

    diffs = result["diffs"]
    if not diffs:
        print(f"[{asurite}] OK — all compared fields match")
        return

    print(f"[{asurite}] DIFF — {len(diffs)} field(s) differ:")
    for field, vals in sorted(diffs.items()):
        print(f"  {field}:")
        print(f"    prod:    {vals['prod']!r}")
        print(f"    nonprod: {vals['nonprod']!r}")


def fetch_all_asurities() -> list[str]:
    from utils.database import User, init_db, session_scope
    init_db()
    with session_scope() as db:
        rows = (
            db.query(User.asurite_id)
            .filter(User.asurite_id.isnot(None), User.verified == True)  # noqa: E712
            .all()
        )
    asurities = []
    for (asurite,) in rows:
        if asurite:
            normalized = asurite.strip()
            if normalized.lower().endswith("@asu.edu"):
                normalized = normalized[: -len("@asu.edu")]
            if normalized:
                asurities.append(normalized)
    return asurities


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare prod vs nonprod Salesforce responses.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("asurities", nargs="*", metavar="ASURITE", help="One or more ASURITE IDs to check.")
    group.add_argument("--all", action="store_true", help="Check every verified ASURITE in the database.")
    parser.add_argument("--workers", type=int, default=20, help="Concurrent workers (default: 20).")
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / "data" / "sf_comparison.xlsx"),
        help="Path for the Excel output file.",
    )
    return parser.parse_args()


def write_xlsx(results: list[dict], output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    GREY   = PatternFill("solid", fgColor="D9D9D9")

    wb = Workbook()
    ws = wb.active
    ws.title = "SF Comparison"

    # ── Header row ──────────────────────────────────────────────────────────
    header_font = Font(bold=True)
    fixed_headers = ["asurite", "status"]
    field_headers = []
    for f in COMPARED_FIELDS:
        field_headers += [f"{f} (prod)", f"{f} (nonprod)"]

    for col, title in enumerate(fixed_headers + field_headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = GREY
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "C2"  # freeze asurite + status columns and header row

    # ── Data rows ───────────────────────────────────────────────────────────
    for row_idx, result in enumerate(results, start=2):
        asurite = result["asurite"]
        prod_err = result.get("prod_err")
        nonprod_err = result.get("nonprod_err")
        prod_found = result["prod_found"]
        nonprod_found = result["nonprod_found"]
        diffs = result.get("diffs", {})

        # Status cell
        if prod_err:
            status = f"PROD ERROR"
        elif nonprod_err:
            status = f"NONPROD ERROR"
        elif not prod_found and not nonprod_found:
            status = "NOT FOUND (either)"
        elif not prod_found:
            status = "NOT FOUND (prod)"
        elif not nonprod_found:
            status = "NOT FOUND (nonprod)"
        elif diffs:
            non_email = {k: v for k, v in diffs.items() if k != "email"}
            status = f"DIFF ({len(non_email)} field{'s' if len(non_email) != 1 else ''})" if non_email else "EMAIL ONLY"
        else:
            status = "OK"

        ws.cell(row=row_idx, column=1, value=asurite)
        ws.cell(row=row_idx, column=2, value=status)

        prod_data = _profile_to_dict(result.get("_prod_profile"))
        nonprod_data = _profile_to_dict(result.get("_nonprod_profile"))

        for field_idx, field in enumerate(COMPARED_FIELDS):
            prod_col = 3 + field_idx * 2
            nonprod_col = prod_col + 1

            prod_val = prod_data.get(field)
            nonprod_val = nonprod_data.get(field)

            prod_cell = ws.cell(row=row_idx, column=prod_col, value=str(prod_val) if prod_val is not None else "")
            nonprod_cell = ws.cell(row=row_idx, column=nonprod_col, value=str(nonprod_val) if nonprod_val is not None else "")

            if not nonprod_found or (prod_val is not None and nonprod_val is None):
                fill = YELLOW
            elif prod_val == nonprod_val:
                fill = GREEN
            else:
                fill = RED

            prod_cell.fill = fill
            nonprod_cell.fill = fill

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18  # asurite
    ws.column_dimensions["B"].width = 22  # status
    for i in range(len(COMPARED_FIELDS) * 2):
        col_letter = ws.cell(row=1, column=3 + i).column_letter
        ws.column_dimensions[col_letter].width = 24

    wb.save(output_path)
    logger.info("Wrote %s", output_path)


def main() -> None:
    args = parse_args()

    apply_env(load_env(ENV_PATH))

    # Re-import after env is loaded so module-level constants pick up the values.
    import asu_discord.salesforce as sf

    prod_creds = {
        "client_id": sf.CLIENT_ID,
        "client_secret": sf.CLIENT_SECRET,
        "base_url": sf.BASE,
    }
    nonprod_creds = {
        "client_id": sf.NONPROD_CLIENT_ID,
        "client_secret": sf.NONPROD_CLIENT_SECRET,
        "base_url": sf.NONPROD_BASE,
        "scope_base_url": sf.NONPROD_SCOPE_BASE,
    }

    missing = [k for k, v in {**prod_creds, **nonprod_creds}.items() if not v]
    if missing:
        sys.exit(f"Missing credentials/config: {missing}. Check your .env file.")

    asurities: list[str] = args.asurities if not args.all else fetch_all_asurities()
    if not asurities:
        sys.exit("No ASURITEs provided.")

    logger.info(
        "Comparing %d ASURITE(s): prod=%s (Basic Auth)  nonprod=%s (OAuth2)",
        len(asurities),
        ESB_BASE,
        sf.NONPROD_BASE,
    )

    total = len(asurities)
    matched = 0
    diverged = 0
    missing_count = 0
    errors = 0
    all_results: list[dict] = []

    import concurrent.futures
    import threading
    print_lock = threading.Lock()

    def process(asurite: str) -> dict:
        return compare_asurite(asurite, prod_creds, nonprod_creds)

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as executor:
        for result in executor.map(process, asurities):
            all_results.append(result)
            with print_lock:
                print_result(result)
                sys.stdout.flush()

            if result.get("prod_err") or result.get("nonprod_err"):
                errors += 1
            elif not result["prod_found"] or not result["nonprod_found"]:
                missing_count += 1
            elif result["diffs"]:
                diverged += 1
            else:
                matched += 1

    print()
    print(f"Summary: {total} checked — {matched} match, {diverged} diverge, {missing_count} not found, {errors} errors")

    write_xlsx(all_results, args.output)
    print(f"Excel report: {args.output}")


if __name__ == "__main__":
    main()
